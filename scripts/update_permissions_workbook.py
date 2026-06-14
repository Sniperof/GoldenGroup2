from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

WORKBOOK = Path('صلاحيات_النظام.xlsx')
MASTER_SHEET = 'كل الصلاحيات'
TEST_SHEET = 'اختبار الأدوار'

HEADER_FILL = PatternFill('solid', fgColor='1F4E78')
HEADER_FONT = Font(color='FFFFFF', bold=True)
GREEN_FILL = PatternFill('solid', fgColor='C6EFCE')
GREEN_FONT = Font(color='006100', bold=True)
RED_FILL = PatternFill('solid', fgColor='FFC7CE')
RED_FONT = Font(color='9C0006', bold=True)
BLUE_FILL = PatternFill('solid', fgColor='D9EAF7')
ORANGE_FILL = PatternFill('solid', fgColor='FCE4D6')
TEAL_FILL = PatternFill('solid', fgColor='E2F0D9')


def set_header(cell, text, fill=HEADER_FILL):
    cell.value = text
    cell.fill = fill
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def copy_row_style(src_ws, dst_ws, src_row, dst_row, max_col):
    for c in range(1, max_col + 1):
        src = src_ws.cell(src_row, c)
        dst = dst_ws.cell(dst_row, c)
        if src.has_style:
            dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


wb = load_workbook(WORKBOOK)
ws = wb[MASTER_SHEET]

# Tracking columns for change control.
if ws.cell(1, 13).value != 'الحالة':
    ws.cell(1, 13).value = 'الحالة'
    ws.cell(1, 14).value = 'ملاحظة التغيير'
    set_header(ws.cell(1, 13), 'الحالة')
    set_header(ws.cell(1, 14), 'ملاحظة التغيير')

# Make sure the new permission exists in the master sheet.
existing_row = None
for r in range(2, ws.max_row + 1):
    if ws.cell(r, 8).value == 'departments.manage':
        existing_row = r
        break

if existing_row is None:
    insert_at = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 8).value == 'departments.view_list':
            insert_at = r + 1
            break
    if insert_at is None:
        insert_at = ws.max_row + 1

    ws.insert_rows(insert_at)

    # Re-number the first column after the insert.
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 1).value = r - 1

    row = insert_at
    values = [
        row - 1,
        999,
        'إدارة النظام والصلاحيات',
        'departments',
        'departments',
        'manage',
        'إدارة الأقسام',
        'departments.manage',
        'عام (كل الفروع) / الفرع',
        11,
        '❌',
        None,
        'مضاف',
        'تمت إضافتها في خطوة الأقسام',
    ]
    for c, value in enumerate(values, start=1):
        ws.cell(row, c).value = value

    for c in range(1, 15):
        ws.cell(row, c).fill = GREEN_FILL
        if c != 11:
            ws.cell(row, c).font = GREEN_FONT

else:
    ws.cell(existing_row, 13).value = 'مضاف'
    ws.cell(existing_row, 14).value = 'تمت إضافتها في خطوة الأقسام'
    for c in range(1, 15):
        ws.cell(existing_row, c).fill = GREEN_FILL
        if c != 11:
            ws.cell(existing_row, c).font = GREEN_FONT

for r in range(2, ws.max_row + 1):
    if ws.cell(r, 13).value is None:
        ws.cell(r, 13).value = 'ثابت'
    if ws.cell(r, 14).value is None:
        ws.cell(r, 14).value = ''

# Build a clean test sheet for role/scope testing.
if TEST_SHEET in wb.sheetnames:
    del wb[TEST_SHEET]
test = wb.create_sheet(TEST_SHEET)

headers = [
    '#', 'المعرف', 'اسم الصلاحية', 'الوحدة', 'الإجراء', 'الحالة',
    'مدير الفرع - الصلاحية', 'مدير الفرع - scope',
    'المشرفة - الصلاحية', 'المشرفة - scope',
    'الفني - الصلاحية', 'الفني - scope',
    'ملاحظات',
]

for c, header in enumerate(headers, start=1):
    fill = HEADER_FILL
    if c in (7, 8):
        fill = BLUE_FILL
    elif c in (9, 10):
        fill = TEAL_FILL
    elif c in (11, 12):
        fill = ORANGE_FILL
    elif c == 6:
        fill = RED_FILL
    elif c == 13:
        fill = GREEN_FILL
    set_header(test.cell(1, c), header, fill)

for r in range(2, ws.max_row + 1):
    values = [
        ws.cell(r, 1).value,
        ws.cell(r, 2).value,
        ws.cell(r, 8).value,
        ws.cell(r, 3).value,
        ws.cell(r, 6).value,
        ws.cell(r, 13).value,
        '', '', '', '', '', '',
        ws.cell(r, 14).value,
    ]
    for c, value in enumerate(values, start=1):
        test.cell(r, c).value = value

    # Copy the master style for the first six columns.
    copy_row_style(ws, test, r, r, 6)
    for c in range(7, 13):
        test.cell(r, c).alignment = Alignment(horizontal='center', vertical='center')
    test.cell(r, 13).alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

    if ws.cell(r, 13).value == 'مضاف':
        for c in range(1, 14):
            test.cell(r, c).fill = GREEN_FILL
            if c != 1:
                test.cell(r, c).font = GREEN_FONT
    elif ws.cell(r, 13).value == 'محذوف':
        for c in range(1, 14):
            test.cell(r, c).fill = RED_FILL
            if c != 1:
                test.cell(r, c).font = RED_FONT

test.freeze_panes = 'A2'
ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:N{ws.max_row}'
test.auto_filter.ref = f'A1:M{test.max_row}'

wb.save(WORKBOOK)
print('Workbook updated')
