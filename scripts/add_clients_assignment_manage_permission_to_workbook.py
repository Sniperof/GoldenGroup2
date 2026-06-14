from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "صلاحيات_النظام.xlsx"

KEY = "clients.assignment.manage"
NAME = "إدارة مسؤولي الزبون"
MODULE_KEY = "clients"
MODULE_NAME = "سجلات الزبائن"
ALLOWED_SCOPES = "عام (كل الفروع) / الفرع"

YES = "✅"
NO = "❌"
GLOBAL = "GLOBAL"
BRANCH = "BRANCH"

FILL_HEADER = PatternFill("solid", fgColor="1F4E78")
FILL_ADDED = PatternFill("solid", fgColor="C6EFCE")
FILL_NO = PatternFill("solid", fgColor="FFC7CE")
FILL_BRANCH = PatternFill("solid", fgColor="D9EAF7")
FILL_GLOBAL = PatternFill("solid", fgColor="D9E1F2")
FILL_NOTE = PatternFill("solid", fgColor="FFF2CC")
FILL_CLIENT = PatternFill("solid", fgColor="EAF2F8")

THIN = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def headers(ws):
    return {cell.value: index for index, cell in enumerate(ws[1], start=1)}


def style_cell(cell, fill=None):
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER
    if fill is not None:
        cell.fill = fill


def style_header(row):
    for cell in row:
        style_cell(cell, FILL_HEADER)
        cell.font = Font(color="FFFFFF", bold=True)


def find_row_by_key(ws, key_col):
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=key_col).value == KEY:
            return row
    return None


def update_all_permissions(wb):
    ws = wb["كل الصلاحيات"]
    h = headers(ws)
    row = find_row_by_key(ws, h["مفتاح الصلاحية"])
    if row is None:
        row = ws.max_row + 1
        ws.cell(row=row, column=h["مفتاح الوحدة"], value=MODULE_KEY)
        ws.cell(row=row, column=h["الوحدة"], value=MODULE_NAME)
        ws.cell(row=row, column=h["مفتاح الصلاحية"], value=KEY)
        ws.cell(row=row, column=h["اسم الصلاحية"], value=NAME)

    if "النطاقات المسموحة" in h:
        ws.cell(row=row, column=h["النطاقات المسموحة"], value=ALLOWED_SCOPES)
    if "الحالة" in h:
        ws.cell(row=row, column=h["الحالة"], value="تمت الإضافة")
    if "ملاحظة التغيير" in h:
        ws.cell(
            row=row,
            column=h["ملاحظة التغيير"],
            value="إضافة جديدة: فصل إدارة مسؤولي الزبون عن تعديل بيانات الزبون. لا تدعم ASSIGNED.",
        )

    for col in range(1, ws.max_column + 1):
        style_cell(ws.cell(row=row, column=col), FILL_ADDED)


def ensure_role_test_columns(ws):
    wanted = [
        "المشرف - الصلاحية",
        "المشرف - scope",
        "موظف إداري - الصلاحية",
        "موظف إداري - scope",
        "مدير الشركة - الصلاحية",
        "مدير الشركة - scope",
        "قرار baseline الزبائن",
    ]
    h = headers(ws)
    col = ws.max_column + 1
    for title in wanted:
        if title not in h:
            ws.cell(row=1, column=col, value=title)
            style_header([ws.cell(row=1, column=col)])
            h[title] = col
            col += 1
    return h


def update_role_test_sheet(wb):
    ws = wb["اختبار الأدوار"]
    h = ensure_role_test_columns(ws)
    row = find_row_by_key(ws, h["اسم الصلاحية"])
    if row is None:
        row = ws.max_row + 1
        ws.cell(row=row, column=h["الوحدة"], value=MODULE_NAME)
        ws.cell(row=row, column=h["اسم الصلاحية"], value=KEY)
        if "الوصف" in h:
            ws.cell(row=row, column=h["الوصف"], value=NAME)

    matrix = {
        "مدير الفرع - الصلاحية": (YES, FILL_ADDED),
        "مدير الفرع - scope": (BRANCH, FILL_BRANCH),
        "المشرفة - الصلاحية": (NO, FILL_NO),
        "المشرفة - scope": ("", FILL_NO),
        "الفني - الصلاحية": (NO, FILL_NO),
        "الفني - scope": ("", FILL_NO),
        "المشرف - الصلاحية": (NO, FILL_NO),
        "المشرف - scope": ("", FILL_NO),
        "موظف إداري - الصلاحية": (YES, FILL_ADDED),
        "موظف إداري - scope": (BRANCH, FILL_BRANCH),
        "مدير الشركة - الصلاحية": (YES, FILL_ADDED),
        "مدير الشركة - scope": (GLOBAL, FILL_GLOBAL),
        "قرار baseline الزبائن": (
            "إضافة جديدة: المشرف لا يدير مسؤولي الزبون؛ الموظف الإداري BRANCH؛ مدير الشركة GLOBAL.",
            FILL_NOTE,
        ),
    }

    for col in range(1, ws.max_column + 1):
        style_cell(ws.cell(row=row, column=col), FILL_CLIENT)
    for title, (value, fill) in matrix.items():
        ws.cell(row=row, column=h[title], value=value)
        style_cell(ws.cell(row=row, column=h[title]), fill)


def update_clients_baseline_sheet(wb):
    ws = wb["اعتماد الزبائن"]
    h = headers(ws)
    row = find_row_by_key(ws, h["مفتاح الصلاحية"])
    if row is None:
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=row - 1)

    values = {
        "مفتاح الصلاحية": KEY,
        "اسم الصلاحية": NAME,
        "النطاقات المعتمدة قبل التنفيذ": ALLOWED_SCOPES,
        "المشرف - الصلاحية": NO,
        "المشرف - scope": "",
        "موظف إداري - الصلاحية": YES,
        "موظف إداري - scope": BRANCH,
        "مدير الشركة - الصلاحية": YES,
        "مدير الشركة - scope": GLOBAL,
        "تعليمات التنفيذ": "إضافة جديدة: من يملك هذه الصلاحية فقط يستطيع إرسال assignmentUserIds. clients.can_be_assigned تبقى أهلية ظهور فقط.",
    }
    for title, value in values.items():
        if title in h:
            ws.cell(row=row, column=h[title], value=value)

    for col in range(1, ws.max_column + 1):
        style_cell(ws.cell(row=row, column=col), FILL_ADDED)
    for title in ["المشرف - الصلاحية", "المشرف - scope"]:
        if title in h:
            style_cell(ws.cell(row=row, column=h[title]), FILL_NO)
    for title, fill in {
        "موظف إداري - scope": FILL_BRANCH,
        "مدير الشركة - scope": FILL_GLOBAL,
        "تعليمات التنفيذ": FILL_NOTE,
    }.items():
        if title in h:
            style_cell(ws.cell(row=row, column=h[title]), fill)


def main():
    wb = load_workbook(WORKBOOK)
    update_all_permissions(wb)
    update_role_test_sheet(wb)
    update_clients_baseline_sheet(wb)
    wb.save(WORKBOOK)
    print(f"updated {WORKBOOK}")


if __name__ == "__main__":
    main()
