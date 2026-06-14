from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "صلاحيات_النظام.xlsx"

YES = "✅"
NO = "❌"
GLOBAL = "GLOBAL"
BRANCH = "BRANCH"
ASSIGNED = "ASSIGNED"

AR_GLOBAL = "عام (كل الفروع)"
AR_BRANCH = "الفرع"
AR_ASSIGNED = "المُسند فقط"
AR_GLOBAL_BRANCH = f"{AR_GLOBAL} / {AR_BRANCH}"
AR_ALL_SCOPES = f"{AR_GLOBAL} / {AR_BRANCH} / {AR_ASSIGNED}"

FILL_HEADER = PatternFill("solid", fgColor="1F4E78")
FILL_CLIENT = PatternFill("solid", fgColor="EAF2F8")
FILL_YES = PatternFill("solid", fgColor="C6EFCE")
FILL_NO = PatternFill("solid", fgColor="FFC7CE")
FILL_ASSIGNED = PatternFill("solid", fgColor="E2F0D9")
FILL_BRANCH = PatternFill("solid", fgColor="D9EAF7")
FILL_GLOBAL = PatternFill("solid", fgColor="D9E1F2")
FILL_CHANGE = PatternFill("solid", fgColor="FCE4D6")
FILL_NOTE = PatternFill("solid", fgColor="FFF2CC")

THIN = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


ASSIGNED_TARGET_KEYS = {
    "clients.view_list",
    "clients.view",
    "clients.edit",
    "clients.delete",
    "clients.account_statement.view",
    "clients.call_log.view",
    "clients.call_log.create",
    "clients.call_log.edit",
    "clients.contact_control.edit",
    "clients.contacts.view",
    "clients.contacts.edit",
    "clients.device_warranties.view",
    "clients.devices.view",
    "clients.network.view",
    "clients.parts_stock.view",
    "clients.pre_offers.view",
    "clients.purchase_history.view",
    "clients.visits.view",
}

SUPERVISOR_ASSIGNED_KEYS = ASSIGNED_TARGET_KEYS - {
    "clients.delete",
}


def style_header(row):
    for cell in row:
        cell.fill = FILL_HEADER
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def style_cell(cell):
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def fill_grant_cells(permission_cell, scope_cell, grant: str | None, scope: str | None):
    permission_cell.value = YES if grant else NO
    permission_cell.fill = FILL_YES if grant else FILL_NO
    style_cell(permission_cell)

    scope_cell.value = scope
    if scope == ASSIGNED:
        scope_cell.fill = FILL_ASSIGNED
    elif scope == BRANCH:
        scope_cell.fill = FILL_BRANCH
    elif scope == GLOBAL:
        scope_cell.fill = FILL_GLOBAL
    else:
        scope_cell.fill = FILL_NO
    style_cell(scope_cell)


def get_headers(ws):
    return {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}


def ensure_columns(ws, headers):
    existing = get_headers(ws)
    next_col = ws.max_column + 1
    for header in headers:
        if header not in existing:
            ws.cell(row=1, column=next_col, value=header)
            style_header([ws.cell(row=1, column=next_col)])
            existing[header] = next_col
            next_col += 1
    return existing


def collect_client_rows(ws):
    headers = get_headers(ws)
    rows = []
    for row in range(2, ws.max_row + 1):
        module_key = ws.cell(row=row, column=headers["مفتاح الوحدة"]).value
        perm_key = ws.cell(row=row, column=headers["مفتاح الصلاحية"]).value
        if module_key == "clients" or (isinstance(perm_key, str) and perm_key.startswith("clients.")):
            rows.append(row)
    return rows


def role_matrix_for_key(key: str):
    supervisor = (True, ASSIGNED) if key in SUPERVISOR_ASSIGNED_KEYS else (False, None)
    admin_employee = (True, BRANCH)
    company_manager = (True, GLOBAL)
    if key == "clients.delete":
        note = "قرار baseline: الحذف ليس للمشرف، حتى لو السجل مسند. الموظف الإداري BRANCH ومدير الشركة GLOBAL."
    elif key in ASSIGNED_TARGET_KEYS:
        note = "يجب تمرير branchId و assignedUserId في policy. المشرف يرى/يعدل السجلات المسندة فقط."
    elif key == "clients.create":
        note = "create لا يدعم ASSIGNED لأنه لا يوجد subject سابق. يمنح للموظف الإداري BRANCH ولمدير الشركة GLOBAL."
    elif key in {"clients.cooldown_unlock", "clients.can_be_assigned"}:
        note = "صلاحية تشغيلية/إدارية على مستوى الفرع وليست صلاحية ASSIGNED للمشرف."
    else:
        note = "مراجعة لاحقة عند تنفيذ policy."
    return supervisor, admin_employee, company_manager, note


def update_all_permissions(wb):
    ws = wb["كل الصلاحيات"]
    headers = get_headers(ws)
    client_rows = collect_client_rows(ws)
    for row in client_rows:
        key = ws.cell(row=row, column=headers["مفتاح الصلاحية"]).value
        allowed_cell = ws.cell(row=row, column=headers["النطاقات المسموحة"])
        status_cell = ws.cell(row=row, column=headers["الحالة"])
        note_cell = ws.cell(row=row, column=headers["ملاحظة التغيير"])
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = FILL_CLIENT
            ws.cell(row=row, column=col).border = BORDER
        if key in ASSIGNED_TARGET_KEYS and allowed_cell.value != AR_ALL_SCOPES:
            allowed_cell.value = AR_ALL_SCOPES
            allowed_cell.fill = FILL_CHANGE
            status_cell.value = "مخطط للتنفيذ"
            status_cell.fill = FILL_CHANGE
            note_cell.value = "تم ضبط baseline الزبائن: إضافة ASSIGNED لدعم صلاحيات المشرف على السجلات المسندة."
            note_cell.fill = FILL_NOTE
        elif key in {"clients.create", "clients.cooldown_unlock", "clients.can_be_assigned"}:
            allowed_cell.value = AR_GLOBAL_BRANCH
            note_cell.value = "Baseline الزبائن: لا يمنح للمشرف ASSIGNED؛ يمنح BRANCH للموظف الإداري و GLOBAL لمدير الشركة."
            note_cell.fill = FILL_NOTE


def update_role_test_sheet(wb):
    ws = wb["اختبار الأدوار"]
    headers = ensure_columns(
        ws,
        [
            "المشرف - الصلاحية",
            "المشرف - scope",
            "موظف إداري - الصلاحية",
            "موظف إداري - scope",
            "مدير الشركة - الصلاحية",
            "مدير الشركة - scope",
            "قرار baseline الزبائن",
        ],
    )
    key_col = headers["اسم الصلاحية"]
    module_col = headers["الوحدة"]
    client_rows = []
    for row in range(2, ws.max_row + 1):
        key = ws.cell(row=row, column=key_col).value
        module = ws.cell(row=row, column=module_col).value
        if module == "سجلات الزبائن" or (isinstance(key, str) and key.startswith("clients.")):
            client_rows.append(row)

    for row in client_rows:
        key = ws.cell(row=row, column=key_col).value
        supervisor, admin_employee, company_manager, note = role_matrix_for_key(key)
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = FILL_CLIENT
            ws.cell(row=row, column=col).border = BORDER

        fill_grant_cells(
            ws.cell(row=row, column=headers["المشرف - الصلاحية"]),
            ws.cell(row=row, column=headers["المشرف - scope"]),
            YES if supervisor[0] else None,
            supervisor[1],
        )
        fill_grant_cells(
            ws.cell(row=row, column=headers["موظف إداري - الصلاحية"]),
            ws.cell(row=row, column=headers["موظف إداري - scope"]),
            YES if admin_employee[0] else None,
            admin_employee[1],
        )
        fill_grant_cells(
            ws.cell(row=row, column=headers["مدير الشركة - الصلاحية"]),
            ws.cell(row=row, column=headers["مدير الشركة - scope"]),
            YES if company_manager[0] else None,
            company_manager[1],
        )
        note_cell = ws.cell(row=row, column=headers["قرار baseline الزبائن"])
        note_cell.value = note
        note_cell.fill = FILL_NOTE
        style_cell(note_cell)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def create_clients_baseline_sheet(wb):
    if "اعتماد الزبائن" in wb.sheetnames:
        del wb["اعتماد الزبائن"]
    ws = wb.create_sheet("اعتماد الزبائن")
    headers = [
        "#",
        "مفتاح الصلاحية",
        "اسم الصلاحية",
        "النطاقات المعتمدة قبل التنفيذ",
        "المشرف - الصلاحية",
        "المشرف - scope",
        "موظف إداري - الصلاحية",
        "موظف إداري - scope",
        "مدير الشركة - الصلاحية",
        "مدير الشركة - scope",
        "تعليمات التنفيذ",
    ]
    ws.append(headers)
    style_header(ws[1])

    all_ws = wb["كل الصلاحيات"]
    all_headers = get_headers(all_ws)
    for index, row in enumerate(collect_client_rows(all_ws), start=1):
        key = all_ws.cell(row=row, column=all_headers["مفتاح الصلاحية"]).value
        name = all_ws.cell(row=row, column=all_headers["اسم الصلاحية"]).value
        allowed = all_ws.cell(row=row, column=all_headers["النطاقات المسموحة"]).value
        supervisor, admin_employee, company_manager, note = role_matrix_for_key(key)
        ws.append([
            index,
            key,
            name,
            allowed,
            YES if supervisor[0] else NO,
            supervisor[1],
            YES if admin_employee[0] else NO,
            admin_employee[1],
            YES if company_manager[0] else NO,
            company_manager[1],
            note,
        ])

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            style_cell(ws.cell(row=row, column=col))
        for col in [5, 7, 9]:
            ws.cell(row=row, column=col).fill = FILL_YES if ws.cell(row=row, column=col).value == YES else FILL_NO
        for col in [6, 8, 10]:
            value = ws.cell(row=row, column=col).value
            ws.cell(row=row, column=col).fill = (
                FILL_ASSIGNED if value == ASSIGNED else
                FILL_BRANCH if value == BRANCH else
                FILL_GLOBAL if value == GLOBAL else
                FILL_NO
            )
        ws.cell(row=row, column=4).fill = FILL_CHANGE if "المُسند" in str(ws.cell(row=row, column=4).value) else FILL_CLIENT
        ws.cell(row=row, column=11).fill = FILL_NOTE

    widths = {
        1: 6,
        2: 34,
        3: 34,
        4: 30,
        5: 18,
        6: 16,
        7: 22,
        8: 18,
        9: 22,
        10: 18,
        11: 72,
    }
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def apply_general_format(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        for cell in ws[1]:
            cell.fill = FILL_HEADER
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER

    # Widen new role-test columns.
    ws = wb["اختبار الأدوار"]
    for col in range(14, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 24 if col != ws.max_column else 72


def main():
    wb = load_workbook(WORKBOOK)
    update_all_permissions(wb)
    update_role_test_sheet(wb)
    create_clients_baseline_sheet(wb)
    apply_general_format(wb)
    wb.save(WORKBOOK)


if __name__ == "__main__":
    main()
