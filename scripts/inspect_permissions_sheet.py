import sys
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')

path = 'صلاحيات_النظام.xlsx'
wb = load_workbook(path)
print(wb.sheetnames)
for ws in wb.worksheets:
    print('\nSHEET:', ws.title, ws.max_row, ws.max_column)
    headers = [ws.cell(1, c).value for c in range(1, min(ws.max_column, 20) + 1)]
    print('HEADERS:', headers)
    for r in range(1, min(ws.max_row, 6) + 1):
        vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 14) + 1)]
        fills = [ws.cell(r, c).fill.fgColor.rgb if ws.cell(r, c).fill and ws.cell(r, c).fill.fgColor.type == 'rgb' else ws.cell(r, c).fill.fgColor.indexed for c in range(1, min(ws.max_column, 14) + 1)]
        print('ROW', r, vals)
        print('FILL', fills)
    for key in ['departments.view_list', 'branches.manage', 'branches.view', 'employees.view_list']:
        found = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 8).value == key:
                found = r
                break
        print('KEY', key, 'ROW', found)
