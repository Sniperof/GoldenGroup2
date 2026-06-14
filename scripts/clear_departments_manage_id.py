from openpyxl import load_workbook

wb = load_workbook('صلاحيات_النظام.xlsx')

for sheet_name in ['كل الصلاحيات', 'اختبار الأدوار']:
    ws = wb[sheet_name]
    for r in range(2, ws.max_row + 1):
        key = ws.cell(r, 8).value if sheet_name == 'كل الصلاحيات' else ws.cell(r, 3).value
        if key == 'departments.manage':
            ws.cell(r, 2).value = None
            break

wb.save('صلاحيات_النظام.xlsx')
print('cleared')
