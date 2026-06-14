import sys
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding='utf-8')

wb = load_workbook('صلاحيات_النظام.xlsx')
print(wb.sheetnames)

ws = wb['كل الصلاحيات']
print('MASTER_HEADERS', [ws.cell(1, c).value for c in range(1, 15)])
for r in range(1, ws.max_row + 1):
    if ws.cell(r, 8).value == 'departments.manage':
        print('DEPT_ROW', r, [ws.cell(r, c).value for c in range(1, 15)])
        break

test = wb['اختبار الأدوار']
print('TEST_HEADERS', [test.cell(1, c).value for c in range(1, 14)])
for r in range(2, min(test.max_row, 4) + 1):
    print('TEST_ROW', r, [test.cell(r, c).value for c in range(1, 14)])
